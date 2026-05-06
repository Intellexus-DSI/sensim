from __future__ import annotations

import logging
import os
from argparse import ArgumentParser, Namespace
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

import numpy as np
import pandas as pd
import torch
from huggingface_hub import login
from tqdm.auto import tqdm

from candidates_builder import calculate_wylie_distance
from candidates_builder.faiss_index import build_faiss_index
from common_utils import load_dataframe, setup_logging, save_dataframe_single, load_yaml
from app_config import AppConfig

LOGGER = logging.getLogger(__name__)

app_config = AppConfig()

from datetime import datetime


def _now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


# -------------------------
# Stats helpers
# -------------------------
def _log_array_stats(name: str, x: np.ndarray) -> None:
    x = np.asarray(x)
    finite = np.isfinite(x)
    n = int(x.size)
    nf = int(finite.sum())
    LOGGER.info("%s stats: N=%d | finite=%d | nonfinite=%d", name, n, nf, n - nf)
    if nf == 0:
        return
    xf = x[finite].astype(np.float64, copy=False)
    qs = np.percentile(xf, [0, 1, 5, 25, 50, 75, 95, 99, 100])
    LOGGER.info(
        "%s percentiles: min=%.6f p01=%.6f p05=%.6f p25=%.6f p50=%.6f p75=%.6f p95=%.6f p99=%.6f max=%.6f",
        name, qs[0], qs[1], qs[2], qs[3], qs[4], qs[5], qs[6], qs[7], qs[8],
    )


def _log_counts(prefix: str, **counts: int) -> None:
    msg = " | ".join([f"{k}={v}" for k, v in counts.items()])
    LOGGER.info("%s: %s", prefix, msg)


# -------------------------
# Memory estimation logs (your code; unchanged)
# -------------------------
def _fmt_bytes(n: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    x = float(n)
    for u in units:
        if x < 1024 or u == units[-1]:
            return f"{x:.2f} {u}"
        x /= 1024
    return f"{x:.2f} TB"


def _log_cuda_mem(prefix: str = "") -> None:
    if not torch.cuda.is_available():
        LOGGER.info("%sCUDA not available.", f"{prefix} " if prefix else "")
        return

    try:
        free_b, total_b = torch.cuda.mem_get_info()
        alloc_b = torch.cuda.memory_allocated()
        reserv_b = torch.cuda.memory_reserved()
        dev = torch.cuda.current_device()
        name = torch.cuda.get_device_name(dev)
        LOGGER.info(
            "%sCUDA mem | device=%d (%s) | free=%s / total=%s | torch allocated=%s | torch reserved=%s",
            f"{prefix} " if prefix else "",
            dev,
            name,
            _fmt_bytes(int(free_b)),
            _fmt_bytes(int(total_b)),
            _fmt_bytes(int(alloc_b)),
            _fmt_bytes(int(reserv_b)),
        )
    except Exception as e:
        LOGGER.warning("%sFailed to query CUDA memory: %s", f"{prefix} " if prefix else "", e)


def _log_faiss_gemm_estimate(
        *,
        n_db: int,
        dim: int,
        query_block: int,
        dtype: np.dtype = np.dtype("float32"),
        workspace_factor: Tuple[float, float] = (0.5, 2.0),
        copies_factor: Tuple[float, float] = (1.0, 1.3),
        prefix: str = "",
) -> None:
    bytes_per = int(np.dtype(dtype).itemsize)

    out_b = int(query_block) * int(n_db) * bytes_per
    xb_b = int(n_db) * int(dim) * bytes_per
    xq_b = int(query_block) * int(dim) * bytes_per

    low = out_b + int(workspace_factor[0] * out_b) + int(copies_factor[0] * (xb_b + xq_b))
    high = out_b + int(workspace_factor[1] * out_b) + int(copies_factor[1] * (xb_b + xq_b))

    LOGGER.info(
        "%sFAISS GEMM est (dtype=%s): qblock=%d, N=%d, D=%d | out=%s | xb=%s | xq=%s | peak≈%s..%s "
        "(includes output + workspace + possible copies)",
        f"{prefix} " if prefix else "",
        np.dtype(dtype).name,
        int(query_block),
        int(n_db),
        int(dim),
        _fmt_bytes(out_b),
        _fmt_bytes(xb_b),
        _fmt_bytes(xq_b),
        _fmt_bytes(low),
        _fmt_bytes(high),
    )


@dataclass
class BuildConfig:
    load_embeddings_from: Optional[Path]
    sentences_file: Optional[Path]
    sentence_col: Optional[str]
    id_col: Optional[str]

    model_name: Optional[str]
    results_dir: Path

    batch_size: int
    seed: int

    top_anchors: int
    final_k: int
    mean_k: int
    pool_k: int
    max_search_k: int

    lev_max_dist: int
    skip_lev_filter: bool

    overlap_k: int
    enforce_global_no_overlap: bool

    faiss_use_gpu: bool


def parse_args(argv: Optional[List[str]] = None) -> Namespace:
    p = ArgumentParser(description="Build labeling dataset for reranker assignment.")
    p.add_argument("--sentences-file", type=str, default=None)
    p.add_argument("--sentence-col", type=str, default=None)
    p.add_argument("--id-col", type=str, default=None)

    p.add_argument("--model-name", type=str, default=None)
    p.add_argument("--results-dir", type=str, default="results/reranker_labelset")

    p.add_argument(
        "--load-embeddings-from",
        type=str,
        default=None,
        help=(
            "Path to .npz with precomputed embeddings. If provided, skips encoding and loads from this file. "
            "The .npz must contain arrays: 'embeddings' (N,D), 'sentence' (N,), 'sent_id' (N,)."
        ),
    )

    p.add_argument("--batch-size", type=int, default=64)
    p.add_argument("--seed", type=int, default=42)

    p.add_argument("--top-anchors", type=int, default=200)
    p.add_argument("--final-k", type=int, default=20)
    p.add_argument("--mean-k", type=int, default=20)
    p.add_argument("--pool-k", type=int, default=100)
    p.add_argument("--max-search-k", type=int, default=100)

    p.add_argument("--lev-max-dist", type=int, default=5)
    p.add_argument(
        "--skip-lev-filter",
        action="store_true",
        help="If set, do NOT drop near-duplicate neighbors using calculate_wylie_distance/lev_max_dist.",
    )

    p.add_argument("--overlap-k", type=int, default=20)
    p.add_argument("--no-global-no-overlap", action="store_true")
    p.add_argument("--faiss-cpu", action="store_true")

    p.add_argument("--no-console-log", action="store_true")
    p.add_argument("--log-dir", type=str, default="./logs")
    p.add_argument("--log-file", type=str, default=Path(__file__).stem + ".log")
    p.add_argument(
        "--log-level",
        type=str,
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
    )
    return p.parse_args(argv)


def encode_sentences(model_name: str, sentences: List[str], batch_size: int) -> np.ndarray:
    from sentence_transformers import SentenceTransformer  # type: ignore

    device = "cuda" if torch.cuda.is_available() else ("mps" if torch.backends.mps.is_available() else "cpu")
    LOGGER.info("Loading SentenceTransformer model=%s device=%s", model_name, device)
    model = SentenceTransformer(model_name, device=device)

    LOGGER.info("Encoding N=%d sentences (batch_size=%d)", len(sentences), batch_size)
    emb = model.encode(
        sentences,
        batch_size=batch_size,
        show_progress_bar=True,
        convert_to_numpy=True,
        normalize_embeddings=False,
    ).astype(np.float32, copy=False)
    return emb


def _load_embeddings_npz(npz_path: Path) -> Tuple[np.ndarray, List[str], List[object]]:
    npz_path = Path(npz_path)
    if not npz_path.exists():
        raise FileNotFoundError(f"Embeddings file not found: {npz_path}")
    if npz_path.suffix.lower() != ".npz":
        raise ValueError(f"--load-embeddings-from must be a .npz file. Got: {npz_path.suffix}")

    data = np.load(npz_path, allow_pickle=True)

    for k in ("embeddings", "sentence", "sent_id"):
        if k not in data:
            raise KeyError(f"Embeddings npz missing required key '{k}'. Keys: {list(data.keys())}")

    embeddings = np.asarray(data["embeddings"], dtype=np.float32)
    if embeddings.ndim != 2:
        raise ValueError(f"'embeddings' must be 2D (N,D). Got shape={embeddings.shape}")

    sentences = [str(s) for s in data["sentence"].tolist()]
    sent_ids = data["sent_id"].tolist()

    if len(sentences) != embeddings.shape[0]:
        raise ValueError(f"Alignment error: embeddings N={embeddings.shape[0]} vs sentence N={len(sentences)}")
    if len(sent_ids) != embeddings.shape[0]:
        raise ValueError(f"Alignment error: embeddings N={embeddings.shape[0]} vs sent_id N={len(sent_ids)}")

    return embeddings, sentences, sent_ids


def _prepare_sentences_and_ids(
        df: pd.DataFrame,
        *,
        sentence_col: str,
        id_col: Optional[str],
        drop_empty: bool = True,
) -> Tuple[List[str], List[object]]:
    """
    Ensures SentenceTransformer gets List[str] only.
    Drops NaNs / empty strings (by default) and keeps ids aligned.
    """
    s = df[sentence_col]

    # detect invalids (NaN, None, non-string that becomes empty after stripping)
    is_na = s.isna()

    # convert everything else to string safely
    s_str = s.where(~is_na, other="").astype(str)

    # normalize whitespace
    s_str = s_str.map(lambda x: x.strip())

    if drop_empty:
        valid_mask = (~is_na) & (s_str != "")
    else:
        # keep empty as "" (SentenceTransformer can encode it, but it’s usually junk)
        valid_mask = ~is_na

    dropped = int((~valid_mask).sum())
    if dropped:
        LOGGER.warning(
            "Dropping %d/%d rows due to empty/NaN/non-text in column '%s'.",
            dropped, len(df), sentence_col
        )

    df2 = df.loc[valid_mask].copy()

    if id_col is not None:
        ids = df2[id_col].tolist()
    else:
        ids = df2.index.tolist()

    sentences = df2[sentence_col].astype(str).map(lambda x: x.strip()).tolist()
    return sentences, ids


def _load_or_encode(cfg: BuildConfig) -> Tuple[np.ndarray, List[str], List[object]]:
    cfg.results_dir.mkdir(parents=True, exist_ok=True)

    if cfg.load_embeddings_from is not None:
        LOGGER.info("Loading embeddings from: %s", cfg.load_embeddings_from)
        embeddings, sentences, ids = _load_embeddings_npz(cfg.load_embeddings_from)
        LOGGER.info("Loaded embeddings: N=%d D=%d", embeddings.shape[0], embeddings.shape[1])
        return embeddings, sentences, ids

    if cfg.sentences_file is None or cfg.sentence_col is None:
        raise ValueError("Encoding mode requires --sentences-file and --sentence-col (or --load-embeddings-from).")
    if cfg.model_name is None:
        raise ValueError("Encoding mode requires --model-name (or --load-embeddings-from).")

    df = load_dataframe(cfg.sentences_file, all_sheets=True)
    if cfg.sentence_col not in df.columns:
        raise KeyError(f"Missing sentence column '{cfg.sentence_col}'. Available: {list(df.columns)}")

    if cfg.id_col is not None and cfg.id_col not in df.columns:
        raise KeyError(f"Missing id column '{cfg.id_col}'. Available: {list(df.columns)}")

    # ✅ sanitize + keep alignment
    sentences, ids = _prepare_sentences_and_ids(
        df,
        sentence_col=cfg.sentence_col,
        id_col=cfg.id_col,
        drop_empty=True,
    )

    LOGGER.info("Prepared N=%d sentences (after cleaning).", len(sentences))

    embeddings = encode_sentences(cfg.model_name, sentences, cfg.batch_size)

    emb_path = cfg.results_dir / "embeddings.npz"
    if emb_path.exists():
        raise FileExistsError(f"Refusing to overwrite existing file: {emb_path}")
    np.savez_compressed(
        emb_path,
        embeddings=embeddings,
        sentence=np.array(sentences, dtype=object),
        sent_id=np.array(ids, dtype=object),
    )
    LOGGER.info("Saved embeddings -> %s", emb_path)

    return embeddings, sentences, ids

def save_labeling_pretty(
    labeling_df: pd.DataFrame,
    filepath: Path,
    *,
    final_k: int,
    blank_rows_between_anchors: int = 3,
    sheet_name: str = "labeling",
    title_prefix: str = "Anchor",
    label_col_name: str = "Label",
) -> None:
    """
    Writes a pretty XLSX labeling file with separated blocks per anchor.

    Output columns: Type | ID | Sentence | Label
    Label column is left empty for manual annotation.

    Expected columns in labeling_df:
      anchor_id, anchor_sentence,
      cand_01_id, cand_01_sentence, ... cand_{final_k:02d}_id, cand_{final_k:02d}_sentence
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    # Validate required columns
    for col in ("anchor_id", "anchor_sentence"):
        if col not in labeling_df.columns:
            raise KeyError(f"labeling_df missing required column: {col}")

    for t in range(1, final_k + 1):
        id_col = f"cand_{t:02d}_id"
        sent_col = f"cand_{t:02d}_sentence"
        if id_col not in labeling_df.columns or sent_col not in labeling_df.columns:
            raise KeyError(f"labeling_df missing columns: {id_col} / {sent_col}")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Column widths: Type | ID | Sentence | Label
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 140
    ws.column_dimensions["D"].width = 18

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    bold = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")
    left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row_cursor = 1

    for idx, r in enumerate(labeling_df.to_dict("records"), start=1):
        # ----- Title row (merged across 4 columns) -----
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=4)
        c = ws.cell(row=row_cursor, column=1, value=f"{title_prefix} {idx}")
        c.font = title_font
        c.alignment = center
        row_cursor += 1

        # ----- Header row -----
        ws.cell(row=row_cursor, column=1, value="Type").font = header_font
        ws.cell(row=row_cursor, column=2, value="ID").font = header_font
        ws.cell(row=row_cursor, column=3, value="Sentence").font = header_font
        ws.cell(row=row_cursor, column=4, value=label_col_name).font = header_font
        for col in range(1, 5):
            ws.cell(row=row_cursor, column=col).alignment = center
        row_cursor += 1

        # ----- Anchor row -----
        ws.cell(row=row_cursor, column=1, value="Anchor").font = bold
        ws.cell(row=row_cursor, column=2, value=r["anchor_id"])
        ws.cell(row=row_cursor, column=3, value=r["anchor_sentence"])
        ws.cell(row=row_cursor, column=4, value="")  # empty label cell
        for col in range(1, 5):
            ws.cell(row=row_cursor, column=col).alignment = left_top_wrap
        row_cursor += 1

        # ----- Candidate rows -----
        for t in range(1, final_k + 1):
            ws.cell(row=row_cursor, column=1, value=f"Cand_{t:03d}").font = bold
            ws.cell(row=row_cursor, column=2, value=r[f"cand_{t:02d}_id"])
            ws.cell(row=row_cursor, column=3, value=r[f"cand_{t:02d}_sentence"])
            ws.cell(row=row_cursor, column=4, value="")  # empty label cell
            for col in range(1, 5):
                ws.cell(row=row_cursor, column=col).alignment = left_top_wrap
            row_cursor += 1

        # ----- Blank separator rows -----
        row_cursor += int(blank_rows_between_anchors)

    wb.save(filepath)


def build_labelset(cfg: BuildConfig) -> Tuple[pd.DataFrame, pd.DataFrame]:
    embeddings, sentences, ids = _load_or_encode(cfg)
    N = len(sentences)
    D = int(embeddings.shape[1])

    LOGGER.info("Working set: N=%d D=%d", N, D)

    # Basic sanity stats
    _log_counts("Config",
                top_anchors=cfg.top_anchors,
                final_k=cfg.final_k,
                mean_k=cfg.mean_k,
                pool_k=cfg.pool_k,
                max_search_k=cfg.max_search_k,
                overlap_k=cfg.overlap_k,
                lev_max_dist=cfg.lev_max_dist)

    # Memory logs
    if cfg.faiss_use_gpu:
        _log_cuda_mem(prefix="Before FAISS index")
        qblock = min(512, N)
        _log_faiss_gemm_estimate(n_db=N, dim=D, query_block=qblock, dtype=np.dtype("float32"), prefix="FAISS search")

    # FAISS search
    K = int(cfg.max_search_k) + 1  # includes self
    if K <= cfg.final_k + 1:
        raise ValueError("--max-search-k must be > --final-k")

    index = build_faiss_index(embeddings, use_gpu=cfg.faiss_use_gpu)

    if cfg.faiss_use_gpu:
        _log_cuda_mem(prefix="After FAISS index build (before search)")

    LOGGER.info("FAISS searching: N=%d, K=%d (includes self). Expecting DISTANCES (lower=closer).", N, K)
    dists, neigh = index.search(embeddings.astype(np.float32, copy=False), K)
    LOGGER.info("FAISS search done.")
    if cfg.faiss_use_gpu:
        _log_cuda_mem(prefix="After FAISS search")

    # quick check: self neighbor distance should be ~0 for L2; can vary by metric/index
    try:
        self_hits = int(np.mean(neigh[:, 0] == np.arange(N)))
        LOGGER.info("Self-hit sanity: neigh[i,0]==i for ~%d/%d (metric dependent).", self_hits, N)
    except Exception:
        pass

    rng = np.random.default_rng(cfg.seed)

    # mean distance (lower is better)
    mean_dist = np.full((N,), np.inf, dtype=np.float32)
    final_neighbors: List[List[int]] = [[] for _ in range(N)]
    final_neighbor_dists: List[List[float]] = [[] for _ in range(N)]

    need_pool = max(cfg.pool_k, cfg.mean_k, cfg.final_k, cfg.overlap_k)
    if cfg.max_search_k < need_pool:
        LOGGER.warning("max_search_k=%d < needed_pool=%d; consider increasing for better coverage.",
                       cfg.max_search_k, need_pool)

    it = tqdm(range(N), desc="Filtering neighbors", unit="sent")
    dropped_not_enough_pool = 0
    for i in it:
        cand_ids = [int(x) for x in neigh[i, 1:].tolist()]  # drop self
        cand_dists = [float(x) for x in dists[i, 1:].tolist()]

        anchor_text = sentences[i]

        filtered_pool: List[int] = []
        filtered_pool_dists: List[float] = []

        # Keep nearest by distance (already sorted by FAISS)
        for nid, nd in zip(cand_ids, cand_dists):
            # drop near-duplicates by your wylie distance
            if not cfg.skip_lev_filter:
                d = calculate_wylie_distance(anchor_text, sentences[nid])
                if d <= cfg.lev_max_dist:
                    continue
            filtered_pool.append(nid)
            filtered_pool_dists.append(nd)
            if len(filtered_pool) >= cfg.pool_k:
                break

        if len(filtered_pool) < cfg.final_k or len(filtered_pool) < cfg.mean_k:
            dropped_not_enough_pool += 1
            continue

        top_ids = filtered_pool[:cfg.final_k]
        top_dists = filtered_pool_dists[:cfg.final_k]

        mk = cfg.mean_k
        mean_dist[i] = float(np.mean(filtered_pool_dists[:mk]))

        final_neighbors[i] = top_ids
        final_neighbor_dists[i] = top_dists

    _log_counts("Neighbor filtering summary",
                total=N,
                usable=int(np.isfinite(mean_dist).sum()),
                dropped_not_enough_pool=dropped_not_enough_pool)

    _log_array_stats("mean_distance (over mean_k nearest after filtering)", mean_dist)

    # Rank anchors by mean distance (ascending: smaller distance = more similar)
    order = np.argsort(mean_dist)

    selected_rows: List[Dict[str, object]] = []
    used_any: set[int] = set()
    used_as_candidate: set[int] = set()

    skipped_overlap = 0
    skipped_no_clean_set = 0

    for i in order:
        if len(selected_rows) >= cfg.top_anchors:
            break
        if not np.isfinite(mean_dist[i]):
            continue
        if i in used_as_candidate:
            skipped_overlap += 1
            continue

        cand = final_neighbors[i]
        cand_dists = final_neighbor_dists[i]
        if len(cand) < cfg.final_k:
            continue

        if cfg.enforce_global_no_overlap:
            kept: List[int] = []
            kept_dists: List[float] = []
            for nid, nd in zip(cand, cand_dists):
                if nid in used_any or nid == i:
                    continue
                kept.append(nid)
                kept_dists.append(nd)
                if len(kept) >= cfg.final_k:
                    break
            if len(kept) < cfg.final_k:
                skipped_no_clean_set += 1
                continue
            cand, cand_dists = kept, kept_dists

        row: Dict[str, object] = {
            "anchor_id": ids[i],
            "anchor_sentence": sentences[i],
            "mean_distance": float(mean_dist[i]),
        }

        for t in range(cfg.final_k):
            row[f"cand_{t + 1:02d}_id"] = ids[cand[t]]
            row[f"cand_{t + 1:02d}_sentence"] = sentences[cand[t]]
            row[f"cand_{t + 1:02d}_distance"] = float(cand_dists[t])

        selected_rows.append(row)

        used_as_candidate.update(cand[:cfg.overlap_k])
        if cfg.enforce_global_no_overlap:
            used_any.add(i)
            used_any.update(cand)

    if len(selected_rows) == 0:
        raise RuntimeError("No anchors selected. Try increasing --max-search-k or relaxing filters.")

    _log_counts("Selection summary",
                selected=int(len(selected_rows)),
                requested=int(cfg.top_anchors),
                skipped_overlap=skipped_overlap,
                skipped_no_clean_set=skipped_no_clean_set)

    detailed_df = pd.DataFrame(selected_rows)

    # Labeling DF: shuffle candidate order per anchor (distance values not needed for labeling)
    label_rows: List[Dict[str, object]] = []
    for _, r in detailed_df.iterrows():
        cand_items = []
        for t in range(cfg.final_k):
            cand_items.append((
                r[f"cand_{t + 1:02d}_id"],
                r[f"cand_{t + 1:02d}_sentence"],
            ))
        rng.shuffle(cand_items)

        out: Dict[str, object] = {
            "anchor_id": r["anchor_id"],
            "anchor_sentence": r["anchor_sentence"],
        }
        for t, (cid, csent) in enumerate(cand_items, start=1):
            out[f"cand_{t:02d}_id"] = cid
            out[f"cand_{t:02d}_sentence"] = csent

        label_rows.append(out)

    labeling_df = pd.DataFrame(label_rows)
    return detailed_df, labeling_df


def _validate_args_mode(cfg: BuildConfig) -> None:
    if cfg.load_embeddings_from is not None:
        if cfg.model_name is not None:
            LOGGER.warning("Ignoring --model-name because --load-embeddings-from is provided.")
        if cfg.sentences_file is not None or cfg.sentence_col is not None:
            LOGGER.warning("Ignoring --sentences-file/--sentence-col because --load-embeddings-from is provided.")
    else:
        if cfg.sentences_file is None:
            raise ValueError("Encoding mode requires --sentences-file (or provide --load-embeddings-from).")
        if cfg.sentence_col is None:
            raise ValueError("Encoding mode requires --sentence-col (or provide --load-embeddings-from).")
        if cfg.model_name is None:
            raise ValueError("Encoding mode requires --model-name (or provide --load-embeddings-from).")

    if cfg.pool_k < max(cfg.final_k, cfg.mean_k):
        raise ValueError("--pool-k must be >= max(--final-k, --mean-k)")
    if cfg.max_search_k < cfg.pool_k:
        raise ValueError("--max-search-k must be >= --pool-k")


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)

    base_dir = app_config.get('sensim_base_dir', '.')
    log_dir = app_config.get('log_dir', os.path.join(base_dir, 'logs'))

    setup_logging(
        log_dir=log_dir,
        log_file=args.log_file,
        log_level=args.log_level,
        console_enabled=(not args.no_console_log),
        name=__name__,
    )

    base_results_dir = Path(app_config.get('reranker_results_dir', os.path.join(base_dir, 'results', 'reranker_labelset')))
    run_results_dir = base_results_dir / _now_stamp()

    cfg = BuildConfig(
        load_embeddings_from=Path(args.load_embeddings_from) if args.load_embeddings_from else None,
        sentences_file=Path(args.sentences_file) if args.sentences_file else None,
        sentence_col=args.sentence_col,
        id_col=args.id_col,
        model_name=args.model_name,
        batch_size=int(args.batch_size),
        results_dir=run_results_dir,
        seed=int(args.seed),
        top_anchors=int(args.top_anchors),
        final_k=int(args.final_k),
        mean_k=int(args.mean_k),
        pool_k=int(args.pool_k),
        max_search_k=int(args.max_search_k),
        lev_max_dist=int(args.lev_max_dist),
        skip_lev_filter=bool(args.skip_lev_filter),
        overlap_k=int(args.overlap_k),
        enforce_global_no_overlap=(not args.no_global_no_overlap),
        faiss_use_gpu=(not args.faiss_cpu),
    )

    _validate_args_mode(cfg)

    _keys = load_yaml("keys.yaml")
    _hf_token = _keys.get("HF_TOKEN", None)
    if _hf_token:
        try:
            login(token=_hf_token)
            LOGGER.info("Logged into HuggingFace Hub successfully.")
        except Exception as e:
            LOGGER.warning("Failed to log into HuggingFace Hub (continuing): %s", e)
    else:
        LOGGER.info("No HF_TOKEN found; continuing without hub auth.")

    detailed_df, labeling_df = build_labelset(cfg)

    detailed_path = cfg.results_dir / "detailed.xlsx"
    labeling_path = cfg.results_dir / "labeling.xlsx"

    save_dataframe_single(detailed_df, detailed_path)

    save_labeling_pretty(
        labeling_df,
        labeling_path,
        final_k=cfg.final_k,
        blank_rows_between_anchors=3,
    )

    LOGGER.info("Saved:\n- %s\n- %s", detailed_path, labeling_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
