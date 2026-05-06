from __future__ import annotations

import logging
import sys
from argparse import ArgumentParser, Namespace
from typing import List, Optional, Tuple
from pathlib import Path
import hashlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from common_utils import install_global_exception_logging, setup_logging, load_dataframe

from llms import LLMSentencesGenerator

LOGGER = logging.getLogger(__name__)


def _candidate_uid(anchor_sentence: str, r: dict) -> str:
    """
    Stable identifier for a candidate row, so we can preserve the same Cand__{i}_{anchor_id}
    IDs across shuffled + ordered variants.
    """
    parts = [
        str(anchor_sentence or ""),
        str(r.get("model", "") or ""),
        str(r.get("temperature", "") or ""),
        str(r.get("sim_level", "") or ""),
        str("" if pd.isna(r.get("candidate")) else r.get("candidate")),
    ]
    s = "\u241E".join(parts)
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def build_anchor_dfs(
        results_df: pd.DataFrame,
        *,
        shuffle_seed: Optional[int] = 1337,
) -> Tuple[List[pd.DataFrame], List[pd.DataFrame]]:
    """
    Returns (shuffled_dfs, ordered_dfs).

    - shuffled_dfs: per-anchor DataFrames with columns [ID, model, sentence]
      * each anchor's candidates are shuffled independently
      * IDs are assigned by shuffled order: Cand__{i}_{anchor_id}

    - ordered_dfs: per-anchor DataFrames with columns [ID, model, temperature, sim_level, sentence]
      * each anchor's candidates are sorted by sim_level (then model, temperature)
      * IDs are preserved from shuffled_dfs via stable UID mapping (same Cand__{i}_{anchor_id})
    """
    required = {"ID", "anchor_sentence", "candidate", "model", "temperature", "sim_level"}
    missing = required - set(results_df.columns)
    if missing:
        raise KeyError(f"results_df missing required columns: {sorted(missing)}")

    ordered_anchors = (
        results_df[["ID", "anchor_sentence"]]
        .drop_duplicates(subset=["anchor_sentence"], keep="first")
        .reset_index(drop=True)
    )

    shuffled_dfs: List[pd.DataFrame] = []
    ordered_dfs: List[pd.DataFrame] = []

    for i, a in enumerate(ordered_anchors.to_dict("records")):
        anchor_id = str(a["ID"])
        anchor_sentence = str(a["anchor_sentence"])

        g = results_df[results_df["anchor_sentence"] == anchor_sentence].copy()

        # --- shuffled: assign Cand__{i}_{anchor_id} in shuffled order ---
        g_shuf = g.sample(frac=1.0, random_state=(shuffle_seed + i)).reset_index(drop=True)

        uid_to_cand_id = {}
        shuf_rows = [{"ID": f"Anchor_{anchor_id}", "model": "-", "sentence": anchor_sentence}]

        for i, r in enumerate(g_shuf.to_dict("records"), start=1):
            uid = _candidate_uid(anchor_sentence, r)
            cand_id = f"Cand__{i}_{anchor_id}"
            uid_to_cand_id[uid] = cand_id

            shuf_rows.append(
                {
                    "ID": cand_id,
                    "model": str(r.get("model", "")),
                    "sentence": "" if pd.isna(r.get("candidate")) else str(r.get("candidate")),
                }
            )

        shuffled_dfs.append(pd.DataFrame(shuf_rows, columns=["ID", "model", "sentence"]))

        # --- ordered: sort by sim_level, but keep the SAME Cand IDs from shuffled mapping ---
        g_ord = g.copy()
        g_ord["sim_level_sort"] = g_ord["sim_level"].astype(float)
        g_ord["model_sort"] = g_ord["model"].astype(str)
        g_ord["temp_sort"] = g_ord["temperature"].astype(float)

        g_ord = (
            g_ord.sort_values(by=["sim_level_sort", "model_sort", "temp_sort"], kind="stable")
            .drop(columns=["sim_level_sort", "model_sort", "temp_sort"])
            .reset_index(drop=True)
        )

        ord_rows = [
            {"ID": f"Anchor_{anchor_id}", "model": "-", "temperature": "-", "sim_level": "-",
             "sentence": anchor_sentence}
        ]

        for r in g_ord.to_dict("records"):
            uid = _candidate_uid(anchor_sentence, r)
            cand_id = uid_to_cand_id.get(uid)
            if cand_id is None:
                cand_id = f"Cand__UNK_{anchor_id}"

            ord_rows.append(
                {
                    "ID": cand_id,
                    "model": str(r.get("model", "")),
                    "temperature": str(r.get("temperature", "")),
                    "sim_level": str(r.get("sim_level", "")),
                    "sentence": "" if pd.isna(r.get("candidate")) else str(r.get("candidate")),
                }
            )

        ordered_dfs.append(pd.DataFrame(ord_rows, columns=["ID", "model", "temperature", "sim_level", "sentence"]))

    return shuffled_dfs, ordered_dfs


def write_pretty_from_anchor_dfs(
        shuffled_dfs: List[pd.DataFrame],
        ordered_dfs: List[pd.DataFrame],
        *,
        output_basepath: Path,
        blank_rows_between_anchors: int = 3,
        sheet_name_shuffled: str = "generated",
        sheet_name_ordered: str = "generated",
        title_prefix: str = "Anchor",
) -> Tuple[Path, Path]:
    """
    Writes two XLSX files:
      - <base>.shuffled.xlsx : columns [ID, model, sentence]
      - <base>.ordered.xlsx  : columns [ID, model, temperature, sim_level, sentence]

    Returns: (ordered_path, shuffled_path)
    """
    output_basepath = Path(output_basepath)
    output_basepath.parent.mkdir(parents=True, exist_ok=True)

    ordered_path = output_basepath.with_name(output_basepath.stem + ".ordered" + output_basepath.suffix)
    shuffled_path = output_basepath.with_name(output_basepath.stem + ".shuffled" + output_basepath.suffix)

    def _write_one(
            dfs: List[pd.DataFrame],
            path: Path,
            *,
            sheet_name: str,
            headers: List[str],
            col_widths: List[int],
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Set column widths
        for col_letter, w in zip(["A", "B", "C", "D", "E"], col_widths):
            ws.column_dimensions[col_letter].width = w

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        bold = Font(bold=True)

        center = Alignment(horizontal="center", vertical="center")
        left_top_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

        row_cursor = 1
        ncols = len(headers)

        for idx, df in enumerate(dfs, start=1):
            # Title row
            ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=ncols)
            c = ws.cell(row=row_cursor, column=1, value=f"---- {title_prefix} {idx} ----")
            c.font = title_font
            c.alignment = center
            row_cursor += 1

            # Header row
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_cursor, column=col_idx, value=h)
                cell.font = header_font
                cell.alignment = center
            row_cursor += 1

            # Data rows
            for r in df.to_dict("records"):
                # First column bold (Anchor_... / Cand__..._ID)
                ws.cell(row=row_cursor, column=1, value=str(r.get(headers[0], ""))).font = bold

                for col_idx, h in enumerate(headers[1:], start=2):
                    ws.cell(row=row_cursor, column=col_idx, value=str(r.get(h, "")))

                for col in range(1, ncols + 1):
                    ws.cell(row=row_cursor, column=col).alignment = left_top_wrap

                row_cursor += 1

            row_cursor += int(blank_rows_between_anchors)

        wb.save(path)

    # Shuffled (ID, model, sentence)
    _write_one(
        shuffled_dfs,
        shuffled_path,
        sheet_name=sheet_name_shuffled,
        headers=["ID", "model", "sentence"],
        col_widths=[28, 22, 140, 12, 12],
    )

    # Ordered (ID, model, temperature, sim_level, sentence)
    _write_one(
        ordered_dfs,
        ordered_path,
        sheet_name=sheet_name_ordered,
        headers=["ID", "model", "temperature", "sim_level", "sentence"],
        col_widths=[28, 22, 14, 12, 140],
    )

    return ordered_path, shuffled_path


def save_generator_pretty(
        results_df: pd.DataFrame,
        output_basepath: Path,
        *,
        blank_rows_between_anchors: int = 3,
        sheet_name: str = "generated",
        title_prefix: str = "Anchor",
        shuffle_seed: Optional[int] = 1337,
) -> Tuple[Path, Path]:
    """
    High-level convenience wrapper.

    Builds the per-anchor DataFrames (shuffled + ordered) and writes two styled XLSX files.

    Returns: (ordered_path, shuffled_path)
    """
    shuffled_dfs, ordered_dfs = build_anchor_dfs(results_df, shuffle_seed=shuffle_seed)
    return write_pretty_from_anchor_dfs(
        shuffled_dfs=shuffled_dfs,
        ordered_dfs=ordered_dfs,
        output_basepath=Path(output_basepath),
        blank_rows_between_anchors=blank_rows_between_anchors,
        sheet_name_shuffled=sheet_name,
        sheet_name_ordered=sheet_name,
        title_prefix=title_prefix,
    )


def _build_arg_parser() -> ArgumentParser:
    p = ArgumentParser(description="Run the LLM-based synthetic sentence generation pipeline.")

    p.add_argument("--model-config", type=str, help="Path to model configuration file.")
    g = p.add_mutually_exclusive_group()
    g.add_argument("--load-raw", action="store_true", help="Load raw sentences instead of generating.", default=False)
    g.add_argument("--generate", action="store_true", help="Generate synthetic sentences (default)", default=True)
    g.add_argument("--output-path", type=str, default="./results/generate/rerank.xlsx", help="Output path")
    # -------------------------
    # Logging configuration
    # -------------------------
    p.add_argument("--log-dir", type=str, default="./logs", help="Directory for log files (default: ./logs)")
    p.add_argument("--log-file", type=str, default="generate_synthetic_sentences.log",
                   help="Log filename. If empty, auto-generate one.")
    p.add_argument(
        "--log-level",
        type=str,
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Logging level (default: INFO)",
    )

    # Additional arguments
    p.add_argument("--no-console-log", action="store_true", help="Disable console logging (file only)")
    return p


def _parse_args(argv: Optional[List[str]]) -> Namespace:
    p = _build_arg_parser()
    args, _ = p.parse_known_args(argv)
    return args


def generate_synthetic_sentences(argv: Optional[List[str]] = None) -> None:
    args = _parse_args(argv)

    log_path = setup_logging(
        log_dir=args.log_dir,
        log_file=args.log_file,
        log_level=args.log_level,
        console_enabled=(not args.no_console_log),
        name=__name__,
    )
    install_global_exception_logging(LOGGER)

    LOGGER.info("Starting generation of synthetic sentences with args: %s", args)
    LOGGER.info("Log file: %s", log_path)
    model_config = args.model_config
    load_raw = args.load_raw
    debug = not model_config and not load_raw
    output_path = Path(args.output_path)

    LOGGER.info("Model config: %s", model_config)

    llm_generator = None
    if load_raw:
        LOGGER.info("Loading raw sentences instead of generating.")
        output_raw_path = output_path.with_stem(output_path.stem + "_raw")
        if output_raw_path.exists():
            output_df = load_dataframe(output_raw_path)
        else:
            raise FileNotFoundError(f"Output path {output_raw_path} does not exist for loading raw sentences.")

    else:
        llm_generator = LLMSentencesGenerator(config_path=model_config, debug=debug, argv=argv)

        LOGGER.info("Running LLM generation with model: %s", llm_generator.model)
        output_df = llm_generator.run()

    save_generator_pretty(
        output_df,
        output_path.with_stem(output_path.stem + "_pretty"),
        blank_rows_between_anchors=3,
        sheet_name="generated",
        title_prefix="Anchor",
    )


def main() -> int:
    try:
        generate_synthetic_sentences(sys.argv[1:])
        return 0
    except Exception:
        LOGGER.exception("Fatal error. Exiting.")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
