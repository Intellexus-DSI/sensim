"""
Two-step processing of the merged trainset xlsx:

Step 1 — Add 'type' column:
  - is_unicode_score == False  →  type = "emr"
  - is_unicode_score is None   →  type = "bws"

Step 2 — Normalize EMR scores per batch (identified by 'pair_XX' ID prefix):
  - Renames original 'score' to 'unnormalized_score'
  - Writes min-max normalized value (per batch) into 'score' for EMR rows
  - BWS rows: score kept as-is (already in [0, 1])

Input:  merged_trainset_*.xlsx
Output: merged_trainset_*_normal_normalized.xlsx
"""

import openpyxl
from collections import defaultdict
from pathlib import Path

INPUT_PATH = Path(
    "data/NewDataA-D/merged_trainset_2026-02-27_22-20-37-2026-02-27_12-24-40.xlsx"
)
OUTPUT_PATH = INPUT_PATH.with_stem(INPUT_PATH.stem + "_normal_normalized")


def get_batch(row_id: str) -> str:
    """Extract 'pair_XX' prefix from ID."""
    parts = row_id.split("_")
    return f"{parts[0]}_{parts[1]}"


def add_type_column(ws) -> None:
    """Appends a 'type' column based on is_unicode_score."""
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    unicode_col_idx = header.index("is_unicode_score")

    type_col_idx = ws.max_column + 1
    ws.cell(row=1, column=type_col_idx, value="type")

    for row in ws.iter_rows(min_row=2):
        is_unicode = row[unicode_col_idx].value
        ws.cell(row=row[0].row, column=type_col_idx, value="emr" if is_unicode is False else "bws")


def compute_batch_stats(ws, id_idx: int, score_idx: int, type_idx: int) -> dict:
    """Compute per-batch min/max for EMR rows only."""
    batch_scores = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[type_idx] == "emr":
            batch_scores[get_batch(row[id_idx])].append(row[score_idx])

    stats = {}
    for batch, scores in batch_scores.items():
        b_min, b_max = min(scores), max(scores)
        stats[batch] = {"min": b_min, "max": b_max, "range": b_max - b_min}
        print(f"  {batch}: min={b_min:.6f}  max={b_max:.6f}")

    return stats


def normalize_emr_scores(ws) -> None:
    """Backs up 'score' as 'unnormalized_score' and writes normalized values into 'score'."""
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    id_idx = header.index("ID")
    score_idx = header.index("score")
    type_idx = header.index("type")

    print("Computing per-batch EMR score ranges...")
    batch_stats = compute_batch_stats(ws, id_idx, score_idx, type_idx)

    unnorm_col_idx = ws.max_column + 1
    ws.cell(row=1, column=unnorm_col_idx, value="unnormalized_score")

    score_col_idx = score_idx + 1  # 1-based

    for row in ws.iter_rows(min_row=2):
        score = row[score_idx].value
        row_type = row[type_idx].value

        ws.cell(row=row[0].row, column=unnorm_col_idx, value=score)

        if row_type == "emr":
            stats = batch_stats[get_batch(row[id_idx].value)]
            normalized = (score - stats["min"]) / stats["range"]
        else:
            normalized = score

        ws.cell(row=row[0].row, column=score_col_idx, value=normalized)


def process(input_path: Path, output_path: Path) -> None:
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    add_type_column(ws)
    normalize_emr_scores(ws)

    wb.save(output_path)
    print(f"\nSaved to {output_path}")


if __name__ == "__main__":
    process(INPUT_PATH, OUTPUT_PATH)
